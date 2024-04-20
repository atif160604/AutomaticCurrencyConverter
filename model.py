import openpyxl as xl
import json
import urllib.parse
import urllib3
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from urllib3.poolmanager import PoolManager
import datetime
from Vat import Vat
import sqlite3 as sql
import os
import sys
import shutil


def get_persistent_data_dir(app_name="CurrencyExchangeApp"):
    # Path to the Application Support directory for your app
    app_support_dir = os.path.join(os.path.expanduser('~/Desktop'), app_name)

    # Create the directory if it doesn't exist
    if not os.path.exists(app_support_dir):
        os.makedirs(app_support_dir)

    return app_support_dir


if getattr(sys, 'frozen', False):
    # If the application is run as a bundled executable, use the temporary folder
    base_path = sys._MEIPASS
else:
    # If it's run in a normal Python environment, use the directory of the script
    base_path = os.path.dirname(__file__)
# Construct the path to the database
database_path = os.path.join(base_path, 'exchange.db')
excel_file_path = os.path.join(base_path, 'test.xlsx')
image_path = os.path.join(base_path, 'logo.png')

original_excel_path = os.path.join(base_path, 'test.xlsx')
original_db_path = os.path.join(base_path, 'exchange.db')
original_image_path = os.path.join(base_path, 'logo.png')

# Paths to the persistent data files in the Application Support directory
persistent_data_dir = get_persistent_data_dir()
persistent_excel_path = os.path.join(persistent_data_dir, 'test.xlsx')
persistent_db_path = os.path.join(persistent_data_dir, 'exchange.db')
persistent_image_path = os.path.join(persistent_data_dir, 'logo.png')

# Copy the original files to the persistent location if they don't exist
# This is done to ensure that the application starts with your provided data files on the first run
if not os.path.exists(persistent_excel_path):
    shutil.copy(original_excel_path, persistent_excel_path)
if not os.path.exists(persistent_db_path):
    shutil.copy(original_db_path, persistent_db_path)
if not os.path.exists(persistent_image_path):
    shutil.copy(original_image_path, persistent_image_path)

current_datetime = datetime.datetime.now()
formatted_date = current_datetime.strftime("%Y-%m-%d")


def add_first_date():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    database.execute("INSERT INTO dates VALUES (?, ?, ?, ?)", ("2024-02-09", "2024-02-09", "2024-02-11", "2024-02-11",))

    exchange_database.commit()
    exchange_database.close()


def change_date(prev_date_fta, current_date_fta, prev_date_corp, current_date_corp):  # Corrected variable name typo
    exchange_database = sql.connect(persistent_db_path)
    database = exchange_database.cursor()

    # Corrected the UPDATE statement
    database.execute(
        "UPDATE dates SET prev_date_fta = ?, current_date_fta = ?, prev_date_corp = ?, current_data_corp = ?",
        (prev_date_fta, current_date_fta, prev_date_corp, current_date_corp))

    exchange_database.commit()
    exchange_database.close()


def show_all_dates():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()

    database.execute("SELECT * FROM dates")
    items = database.fetchone()

    exchange_database.commit()
    exchange_database.close()
    return items[0], items[1], items[2], items[3]


# database.execute("""CREATE TABLE dates (
#     prev_date_fta text,
#     current_date_fta text,
#     prev_date_corp text,
#     current_data_corp text
#     )
# """)
#
# database.execute("""CREATE TABLE fta_rates (
#     currency text
#     )
# """)
#


def convert_to_bool(text: str):
    if text.upper() == "TRUE":
        return True
    elif text.upper() == "FALSE":
        return False


# database.execute(f"INSERT INTO fta_rates VALUES('AED')")

test_change = Vat("USD", "INR", False, True, True, True)
test_change2 = Vat("USD", "AUD", True, False, True, False)
test_change3 = Vat("USD", "MXN", False, True, False, True)
test_change4 = Vat("INR", "USD", True, False, False, False)

# ALL CURRENCIES
list_of_all_currencies = ['ARS', 'AUD', 'AZN', 'BDT', 'BGN', 'BHD', 'BND', 'BRL', 'BWP', 'BYN', 'CAD', 'CHF', 'CLP',
                          'CNH', 'CNY', 'COP', 'CZK', 'DKK', 'DZD', 'EGP', 'ETB', 'EUR', 'GBP', 'HKD', 'HRK', 'HUF',
                          'IDR', 'ILS', 'INR', 'IQD', 'ISK', 'JOD', 'JPY', 'KES', 'KRW', 'KWD', 'KZT', 'LBP', 'LKR',
                          'LYD', 'MAD', 'MKD', 'MUR', 'MXN', 'MYR', 'NGN', 'NOK', 'NZD', 'OMR', 'PEN', 'PHP', 'PKR',
                          'PLN', 'QAR', 'RON', 'RSD', 'RUB', 'SAR', 'SDG', 'SEK', 'SGD', 'SYP', 'THB', 'TMT', 'TND',
                          'TRY', 'TTD', 'TWD', 'TZS', 'UGX', 'USD', 'UZS', 'VND', 'YER', 'ZAR', 'ZMW']

list_of_all_currencies_Vat = ['AED', 'AFN', 'ALL', 'ARS', 'AUD', 'BAM', 'BDT', 'BGN', 'BHD', 'BND', 'BRL', 'CAD', 'CHF',
                              'CNY', 'CUP', 'CZK', 'DJF', 'DKK', 'DZD', 'EGP', 'ETB', 'EUR', 'GBP', 'GNF', 'HKD', 'HUF',
                              'IDR', 'INR', 'ISK', 'JOD', 'JPY', 'KES', 'KRW', 'KWD', 'LBP', 'LKR', 'LYD', 'MAD', 'MGA',
                              'MMK', 'MUR', 'MXN', 'MYR', 'NGN', 'NOK', 'NZD', 'OMR', 'PHP', 'PKR', 'PLN', 'QAR', 'RON',
                              'RUB', 'SEK', 'SGD', 'THB', 'TJS', 'TND', 'TRY', 'TWD', 'TZS', 'UGX', 'USD', 'VND', 'XAF',
                              'XDR', 'XOF', 'YER', 'ZAR']

# Accessing The API
API_BASE_ADDRESS = 'https://fx-api.fluentax.com'

testEndpoint = "https://fx-api.fluentax.com/v1/Currencies"
endpoint = "https://sso.fluentax.com/auth/realms/fluentax/protocol/openid-connect/token"
client_id = "53d8d57a-8fce-45df-b7f2-e7d20eefe548"
client_secret = "ApvcYs8npXVC09CVzfuXFt9gLNuZHkk0"

parameters = {
    'grant_type': "client_credentials",
    'client_id': client_id,
    'client_secret': client_secret,
    'scope': 'fx_api',
}

# Indexes to Start For FTA Rates
row_to_start_rates = 9
column_to_start_rates = 2

# List of Currencies For FTA Rates
list_of_currency = []

# Dictionary To Store All Conversion values
dictionary_of_currency = {"AED": 1.000000}
dictionary_of_currency_corp = {"SAR": 1.000000}

# List of Vat Currencies They Want
list_of_vat_currencies = []
# Indexes To Start Vat Currencies
row_to_start_rates_vat = 9


def start():
    temp_prev_date_fta, temp_current_date_fta, temp_prev_date_corp, temp_current_date_corp = show_all_dates()
    change_date(temp_prev_date_fta, formatted_date, temp_prev_date_corp, formatted_date)
    show_all_fta()
    show_all_corp()


def add_all():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    # Ensure each currency in the list is a tuple
    currency_tuples = [(currency,) for currency in list_of_currency]

    database.executemany("INSERT INTO fta_rates VALUES (?)", currency_tuples)

    exchange_database.commit()
    exchange_database.close()


def delete_all_corp():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    # Deleting a record
    database.execute("DELETE from corp_rates_final")

    exchange_database.commit()
    exchange_database.close()


def add_all_corp():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    # Ensure each currency in the list is a tuple
    currency_tuples = [
        (currency.init_currency, currency.currency_to_convert, str(currency.is_vat), str(currency.is_corp),
         str(currency.is_red), str(currency.is_highlighted)) for currency in list_of_vat_currencies]

    database.executemany("INSERT INTO corp_rates_final VALUES (?, ?, ?, ?, ?, ?)", currency_tuples)

    exchange_database.commit()
    exchange_database.close()


def show_all_fta():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    # database.execute("""CREATE TABLE corp_rates_final (
    #     converting_from text,
    #     converting_to text,
    #     is_vat text,
    #     is_corp text,
    #     is_red text,
    #     is_highlight text
    #     )
    # """)

    database.execute("SELECT * FROM fta_rates")
    items = database.fetchall()

    for item in items:
        print(item[0])
        list_of_currency.append(item[0])

    exchange_database.commit()
    exchange_database.close()


def show_all_corp():
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()

    database.execute("SELECT * FROM corp_rates_final")
    items = database.fetchall()

    for item in items:
        temp_cur = Vat(item[0], item[1], convert_to_bool(item[2]), convert_to_bool(item[3]), convert_to_bool(item[4]),
                       convert_to_bool(item[5]))
        print(temp_cur)
        list_of_vat_currencies.append(temp_cur)

    exchange_database.commit()
    exchange_database.close()


def add_record_fta(currency):
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    database.execute("INSERT INTO fta_rates VALUES (?)", (currency,))

    exchange_database.commit()
    exchange_database.close()


def delete_record_fta(name):
    exchange_database = sql.connect(persistent_db_path)
    database = exchange_database.cursor()

    database.execute("DELETE from fta_rates WHERE currency = (?)", (name,))
    print("deleted name")
    print(list_of_currency)
    exchange_database.commit()
    exchange_database.close()


def add_record_corp(converting_from, converting_to, is_vat, is_corp, is_red, is_highlight):
    exchange_database = sql.connect(persistent_db_path)

    database = exchange_database.cursor()
    database.execute("INSERT INTO corp_rates_final VALUES (?, ?, ?, ?, ?, ?)",
                     (converting_from, converting_to, is_vat, is_corp, is_red, is_highlight))

    exchange_database.commit()
    exchange_database.close()


def delete_record_corp(converting_from, converting_to, is_vat, is_corp, is_red, is_highlight):
    exchange_database = sql.connect(persistent_db_path)
    database = exchange_database.cursor()

    # Corrected the parameters to be a tuple
    database.execute("""
            DELETE FROM corp_rates_final 
            WHERE converting_from = ? 
            AND converting_to = ? 
            AND is_vat = ? 
            AND is_corp = ? 
            AND is_red = ? 
            AND is_highlight = ? 
        """, (converting_from, converting_to, is_vat, is_corp, is_red, is_highlight))

    exchange_database.commit()
    exchange_database.close()


# add_all()
# add_all_corp()
# add_first_date()
print(show_all_dates())

print(show_all_dates())
start()
# delete_all_corp()
column_to_start_rates_vat = len(list_of_currency) + 2
print(column_to_start_rates_vat)
# Loading The Sheet
excel = xl.load_workbook(persistent_excel_path)
exchange_rate_sheet = excel["Sheet1"]

# Formatting Data
hex_color = "CBDFB8"  # Green
bold_font = Font(bold=True)
calibri_font = Font(name='Calibri', size=11, bold=True)
red_font_bold = red_font = Font(name='Calibri', size=11, color='FF0000', bold=True)
calibri_font_no_bold = Font(name='Calibri', size=11)
thin_side = Side(border_style="thin", color="000000")
thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
center_alignment = Alignment(horizontal='center')
yellow_fill = PatternFill(start_color='FFFF00',
                          end_color='FFFF00',
                          fill_type='solid')


def get_access_token(http: PoolManager) -> str:
    fields: dict[str, str] = {'grant_type': 'client_credentials', 'client_id': client_id,
                              'client_secret': client_secret, 'scope': 'fx_api'}

    token_response = http.request(
        'POST',
        endpoint,
        fields=fields,
        encode_multipart=False
    )
    token_response_data = json.loads(token_response.data.decode('utf-8'))
    access_token: str = token_response_data['access_token']

    return access_token


def retrieve_latest_exchange_rates():
    http = urllib3.PoolManager()

    access_token = get_access_token(http)

    bank_id_aecb = 'AECB'
    bank_id_sacb = 'SACB'
    format = 'json'
    # ----------------------------CENTRAL BANK OF UAE---------------
    # Accessing rates from central bank of UAE
    latest_rates_request_url = urllib.parse.urljoin(
        API_BASE_ADDRESS, f'v1/Banks/{bank_id_aecb}/DailyRates/Latest?format={format}')

    # Rates for central bank of UAE
    latest_rates_response = http.request(
        'GET',
        latest_rates_request_url,
        headers={'Authorization': f'bearer {access_token}'}
    )
    rates = latest_rates_response.json()
    print(rates)

    # ----------------------------SAUDI CENTRAL BANK---------------
    # Accessing rates from saudi central bank
    latest_rates_request_url_corp = urllib.parse.urljoin(
        API_BASE_ADDRESS, f'v1/Banks/{bank_id_sacb}/DailyRates/Latest?format={format}')
    latest_rates_response_corp = http.request(
        'GET',
        latest_rates_request_url_corp,
        headers={'Authorization': f'bearer {access_token}'}
    )
    rates_corp = latest_rates_response_corp.json()
    print(rates_corp)

    temp_prev_date_fta, temp_current_date_fta, temp_prev_date_corp, temp_current_date_corp = show_all_dates()

    try:
        cur_rates = rates["rates"][temp_current_date_fta]
    except:
        cur_rates = rates["rates"][temp_prev_date_fta]
        date_to_use_fta = temp_prev_date_fta
    else:
        date_to_use_fta = temp_current_date_fta

    try:
        cur_rates_corp = rates_corp["rates"][temp_current_date_corp]
    except:
        cur_rates_corp = rates_corp["rates"][temp_prev_date_corp]
        date_to_use_corp = temp_prev_date_corp
    else:
        date_to_use_corp = temp_current_date_corp

    change_date(date_to_use_fta, formatted_date, date_to_use_corp, formatted_date)

    for cur in cur_rates:
        list_of_all_currencies.append(cur)
        print(cur)

    for cur in cur_rates_corp:
        list_of_all_currencies_Vat.append(cur)
        print(cur)
    print(list_of_all_currencies)
    print(list_of_all_currencies_Vat)

    for currency in list_of_all_currencies:
        if currency != "AED":
            dictionary_of_currency[currency] = rates["rates"][date_to_use_fta][currency]["rate"]
    for currency in list_of_all_currencies_Vat:
        if currency != "SAR":
            dictionary_of_currency_corp[currency] = rates_corp["rates"][date_to_use_corp][currency]["rate"]


# Updates FTA Rates
def update_excel_with_rate_values():
    count = 0
    for i in range(column_to_start_rates, len(list_of_currency) + column_to_start_rates):
        exchange_rate_sheet.cell(row=row_to_start_rates, column=i,
                                 value=dictionary_of_currency[list_of_currency[count]]).font = Font(color="FF0000")
        count += 1


# new Conversion For Vat
def new_conversion(currency_start, currency_convert, is_vat, is_corp, is_red, is_highlighted):
    cur = Vat(currency_start, currency_convert, is_vat, is_corp, is_red, is_highlighted)
    list_of_vat_currencies.append(cur)


# Removing Conversion For Vat
def delete_conversion(currency_start, currency_convert, is_vat, is_corp):
    for currency in list_of_vat_currencies:
        if currency_start == currency.init_currency and currency_convert == currency.currency_to_convert and is_vat == currency.is_vat and is_corp == currency.is_corp:
            list_of_vat_currencies.remove(currency)


# Finding If Conversion Is Already in Vat
def find_conversion(currency):
    for cur in list_of_vat_currencies:
        if cur == currency:
            return True
    return False


# Adds All The Required Data To Vat
def reformat_vat_area():
    print("Function started")
    count = 0

    for currency in list_of_vat_currencies:
        # Set initial currency and currency to convert
        init_cell = exchange_rate_sheet.cell(row=row_to_start_rates_vat + count, column=column_to_start_rates_vat,
                                             value=currency.init_currency)
        convert_cell = exchange_rate_sheet.cell(row=row_to_start_rates_vat + count,
                                                column=column_to_start_rates_vat + 1,
                                                value=currency.currency_to_convert)

        # Initialize cells for VAT, Corp, and Inverse rates
        vat_cell = exchange_rate_sheet.cell(row=row_to_start_rates_vat + count, column=column_to_start_rates_vat + 2)
        corp_cell = exchange_rate_sheet.cell(row=row_to_start_rates_vat + count, column=column_to_start_rates_vat + 3)
        inverse_cell = exchange_rate_sheet.cell(row=row_to_start_rates_vat + count,
                                                column=column_to_start_rates_vat + 4)

        # Determine the rate dictionary and set values based on flags
        if currency.is_vat:
            vat_rate = dictionary_of_currency[currency.init_currency] / dictionary_of_currency[
                currency.currency_to_convert]
            inverse_rate = 1 / vat_rate
            vat_cell.value = vat_rate
            corp_cell.value = "-"
            inverse_cell.value = inverse_rate
        elif currency.is_corp:
            corp_rate = dictionary_of_currency_corp[currency.init_currency] / dictionary_of_currency_corp[
                currency.currency_to_convert]
            inverse_rate = 1 / corp_rate
            corp_cell.value = corp_rate
            vat_cell.value = "-"
            inverse_cell.value = inverse_rate

        # Apply styles based on flags to all relevant cells
        for cell in [init_cell, convert_cell, vat_cell, corp_cell, inverse_cell]:
            cell.font = red_font_bold if currency.is_red else calibri_font_no_bold
            cell.border = thin_border
            if currency.is_highlighted:
                cell.fill = yellow_fill

        print("Row", count + 1, "formatted successfully")
        count += 1

    print("Function completed")


def list_of_cells(start_row, start_column, end_row, end_column):
    list_cells = []
    for row in range(start_row, end_row + 1):
        for column in range(start_column, end_column + 1):
            list_cells.append((row, column))
    return list_cells


def reset():
    # Remove all merged cells
    merged_cells_ranges = list(exchange_rate_sheet.merged_cells.ranges)

    # Unmerge all merged cells
    for merged_cell in merged_cells_ranges:
        exchange_rate_sheet.unmerge_cells(str(merged_cell))

    default_font = Font()
    default_border = Border()
    default_fill = PatternFill(fill_type=None)  # Setting fill_type to None will remove the fill
    default_alignment = Alignment()

    # Loop through each cell, reset style attributes, and clear content
    for row in exchange_rate_sheet.iter_rows():
        for cell in row:
            cell.value = None  # Clear the content of the cell
            cell.font = default_font
            cell.border = default_border
            cell.fill = default_fill
            cell.alignment = default_alignment


def add_name_of_currency():
    list_cells = list_of_cells(8, 2, 8, len(list_of_currency) + 1)
    count = 0
    for row, column in list_cells:
        exchange_rate_sheet.cell(row=row, column=column, value=list_of_currency[count])
        count += 1


def formatting():
    # Reset
    reset()

    # Color of cell
    fill_color = PatternFill(start_color=hex_color, fill_type='solid')

    # FTA Rate
    exchange_rate_sheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(list_of_currency) + 1)
    exchange_rate_sheet.cell(row=6, column=1).fill = fill_color
    exchange_rate_sheet.cell(row=6, column=1, value="FTA RATE").alignment = center_alignment
    exchange_rate_sheet.cell(row=6, column=1, value="FTA RATE").font = calibri_font
    exchange_rate_sheet.cell(row=6, column=1, value="FTA RATE").alignment = center_alignment

    # Currency Rate
    exchange_rate_sheet.merge_cells(start_row=7, start_column=1, end_row=7, end_column=len(list_of_currency) + 1)
    exchange_rate_sheet.cell(row=7, column=1).fill = fill_color
    exchange_rate_sheet.cell(row=7, column=1, value="currency rates").alignment = center_alignment
    exchange_rate_sheet.cell(row=7, column=1, value="currency rates").font = calibri_font
    exchange_rate_sheet.cell(row=7, column=1, value="currency rates").alignment = center_alignment

    # Filling Currency Name Rate Clor & Bold
    list_cells = list_of_cells(8, 1, 8, len(list_of_currency) + 1)
    print(list_cells)
    for row, column in list_cells:
        exchange_rate_sheet.cell(row=row, column=column).fill = fill_color
        exchange_rate_sheet.cell(row=row, column=column).font = calibri_font
        exchange_rate_sheet.cell(row=row, column=column).border = thin_border
        exchange_rate_sheet.cell(row=row, column=column).alignment = center_alignment

    # Insert date
    exchange_rate_sheet.cell(row=8, column=1, value="DATE")
    print("insert")

    # FORMATTING VAT AREA MERGING CELLS
    exchange_rate_sheet.merge_cells(start_row=6, start_column=len(list_of_currency) + 2, end_row=8,
                                    end_column=len(list_of_currency) + 2)

    exchange_rate_sheet.merge_cells(start_row=6, start_column=len(list_of_currency) + 3, end_row=8,
                                    end_column=len(list_of_currency) + 3)

    exchange_rate_sheet.merge_cells(start_row=6, start_column=len(list_of_currency) + 4, end_row=6,
                                    end_column=len(list_of_currency) + 6)

    # Filling in Text Values VAT and adjusting Width
    print(len(list_of_currency))
    exchange_rate_sheet.cell(row=7, column=len(list_of_currency) + 4, value="VAT Rate")
    column_width_to_change = get_column_letter(len(list_of_currency) + 4)
    exchange_rate_sheet.column_dimensions[column_width_to_change].width = 13.5

    exchange_rate_sheet.cell(row=7, column=len(list_of_currency) + 5, value="Corporate")
    column_width_to_change = get_column_letter(len(list_of_currency) + 5)
    exchange_rate_sheet.column_dimensions[column_width_to_change].width = 17.33

    exchange_rate_sheet.cell(row=7, column=len(list_of_currency) + 6, value="Inverse Rate")
    column_width_to_change = get_column_letter(len(list_of_currency) + 6)
    exchange_rate_sheet.column_dimensions[column_width_to_change].width = 15.67

    # Adding COLOR to VAT
    list_cells = list_of_cells(6, len(list_of_currency) + 2, 8, len(list_of_currency) + 6)
    print(list_cells)
    for row, column in list_cells:
        exchange_rate_sheet.cell(row=row, column=column).fill = fill_color
        exchange_rate_sheet.cell(row=row, column=column).font = calibri_font
        exchange_rate_sheet.cell(row=row, column=column).border = thin_border
        exchange_rate_sheet.cell(row=row, column=column).alignment = center_alignment


def refresh_excel():
    reset()
    formatting()
    retrieve_latest_exchange_rates()
    update_excel_with_rate_values()
    add_name_of_currency()
    reformat_vat_area()
    try:
        excel.save(persistent_excel_path)
        print("done")
        print(persistent_excel_path)
    except Exception as e:
        print(f"Error saving Excel file: {e}")
